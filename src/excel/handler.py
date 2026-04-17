from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from src.excel.time_utils import (
    compute_preview_hours,
    duration_to_text,
    normalize_excel_time,
)


MONTH_SHEETS = {
    1: "Jan.",
    2: "Feb.",
    3: "März",
    4: "April",
    5: "Mai",
    6: "Jun.",
    7: "Jul.",
    8: "Aug.",
    9: "Sep.",
    10: "Okt.",
    11: "Nov.",
    12: "Dez.",
}

HOLIDAY_CODE = "S"
BRIDGE_DAY_CODE = "G"


@dataclass
class WriteResult:
    target_date: date
    kind: str
    preview_hours: float | None
    windows_copied: bool
    code: str | None = None
    start_text: str | None = None
    end_text: str | None = None
    breaks_text: str | None = None


class ExcelHandler:
    def __init__(self, workbook_path: str, windows_path: str | None = None) -> None:
        self.workbook_path = Path(workbook_path)
        self.windows_path = Path(windows_path) if windows_path else None

    def get_status(self, target_date: date) -> dict:
        workbook = load_workbook(self.workbook_path)
        try:
            auto_code, auto_reason = self._get_auto_code(workbook, target_date)
            sheet = workbook[self._sheet_name_for_date(target_date)]
            row = self._find_row(sheet, target_date)

            start = normalize_excel_time(sheet[f"C{row}"].value)
            end = normalize_excel_time(sheet[f"D{row}"].value)
            breaks = self._read_breaks(sheet, row)
            special_code = self._manual_special_code(sheet[f"S{row}"].value, auto_code)
            soll = self._get_soll_duration(workbook, target_date.weekday())
            preview = compute_preview_hours(start, end, breaks, special_code, soll)

            return {
                "has_entry": start is not None
                or end is not None
                or special_code is not None,
                "is_auto_skip": auto_code is not None,
                "auto_reason": auto_reason,
                "special_code": special_code,
                "start_text": duration_to_text(start),
                "end_text": duration_to_text(end),
                "breaks_text": self._format_breaks(breaks),
                "preview_hours": preview,
            }
        finally:
            workbook.close()

    def write_special_day(self, target_date: date, code: str) -> WriteResult:
        workbook = load_workbook(self.workbook_path)
        try:
            sheet = workbook[self._sheet_name_for_date(target_date)]
            row = self._find_row(sheet, target_date)
            self._clear_time_cells(sheet, row)
            sheet[f"S{row}"] = code.upper()
            workbook.save(self.workbook_path)
            windows_copied = self._copy_to_windows()
            return WriteResult(
                target_date=target_date,
                kind="special",
                preview_hours=None,
                windows_copied=windows_copied,
                code=code.upper(),
            )
        finally:
            workbook.close()

    def write_workday(
        self,
        target_date: date,
        start: timedelta,
        end: timedelta,
        breaks: list[tuple[timedelta, timedelta]],
    ) -> WriteResult:
        workbook = load_workbook(self.workbook_path)
        try:
            auto_code, auto_reason = self._get_auto_code(workbook, target_date)
            if auto_code is not None:
                raise ValueError(
                    f"{target_date.strftime('%d.%m.%Y')} ist {auto_reason.lower()} und wird nicht als Arbeitstag erfasst."
                )

            sheet = workbook[self._sheet_name_for_date(target_date)]
            row = self._find_row(sheet, target_date)
            self._clear_time_cells(sheet, row)
            sheet[f"C{row}"] = start
            sheet[f"D{row}"] = end
            sheet[f"S{row}"] = None

            break_columns = [("F", "G"), ("H", "I"), ("J", "K")]
            for index, (start_col, end_col) in enumerate(break_columns):
                if index < len(breaks):
                    break_start, break_end = breaks[index]
                    sheet[f"{start_col}{row}"] = break_start
                    sheet[f"{end_col}{row}"] = break_end
                else:
                    sheet[f"{start_col}{row}"] = None
                    sheet[f"{end_col}{row}"] = None

            self._apply_time_number_format(sheet, row)

            soll = self._get_soll_duration(workbook, target_date.weekday())
            preview = compute_preview_hours(start, end, breaks, None, soll)
            workbook.save(self.workbook_path)
            windows_copied = self._copy_to_windows()

            return WriteResult(
                target_date=target_date,
                kind="workday",
                preview_hours=preview,
                windows_copied=windows_copied,
                start_text=duration_to_text(start),
                end_text=duration_to_text(end),
                breaks_text=self._format_breaks(breaks),
            )
        finally:
            workbook.close()

    def _copy_to_windows(self) -> bool:
        if self.windows_path is None:
            return False
        self.windows_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.workbook_path, self.windows_path)
        return True

    def _sheet_name_for_date(self, target_date: date) -> str:
        return MONTH_SHEETS[target_date.month]

    def _find_row(self, sheet, target_date: date) -> int:
        for row in range(8, 46):
            cell_value = sheet[f"A{row}"].value
            if isinstance(cell_value, datetime) and cell_value.date() == target_date:
                return row
        raise ValueError(
            f"Kein Tabellenzeile fuer {target_date.strftime('%d.%m.%Y')} gefunden."
        )

    def _clear_time_cells(self, sheet, row: int) -> None:
        for column in ("C", "D", "F", "G", "H", "I", "J", "K"):
            sheet[f"{column}{row}"] = None

    def _apply_time_number_format(self, sheet, row: int) -> None:
        for column in ("C", "D", "F", "G", "H", "I", "J", "K"):
            sheet[f"{column}{row}"].number_format = "hh:mm"

    def _read_breaks(self, sheet, row: int) -> list[tuple[timedelta, timedelta]]:
        breaks: list[tuple[timedelta, timedelta]] = []
        for start_col, end_col in (("F", "G"), ("H", "I"), ("J", "K")):
            start = normalize_excel_time(sheet[f"{start_col}{row}"].value)
            end = normalize_excel_time(sheet[f"{end_col}{row}"].value)
            if start is not None and end is not None:
                breaks.append((start, end))
        return breaks

    def _format_breaks(self, breaks: list[tuple[timedelta, timedelta]]) -> str:
        return ", ".join(
            f"{duration_to_text(start)}-{duration_to_text(end)}"
            for start, end in breaks
        )

    def _manual_special_code(self, cell_value, auto_code: str | None) -> str | None:
        if not isinstance(cell_value, str):
            return None
        cleaned = cell_value.strip().upper()
        if not cleaned:
            return None
        if cleaned.startswith("="):
            return None
        if auto_code and cleaned == auto_code:
            return None
        return cleaned

    def _get_auto_code(
        self, workbook, target_date: date
    ) -> tuple[str | None, str | None]:
        if target_date.weekday() >= 5:
            return "U", "Wochenende"

        allgemein = workbook["Allgemein"]
        holiday_dates = self._collect_date_set(allgemein, "L", 4, 27)
        bridge_dates = self._collect_date_set(allgemein, "L", 32, 40)

        if target_date in holiday_dates:
            return HOLIDAY_CODE, "Feiertag"
        if target_date in bridge_dates:
            return BRIDGE_DAY_CODE, "Brueckentag"
        return None, None

    def _collect_date_set(
        self, sheet, column: str, start_row: int, end_row: int
    ) -> set[date]:
        values: set[date] = set()
        for row in range(start_row, end_row + 1):
            value = sheet[f"{column}{row}"].value
            if isinstance(value, datetime):
                values.add(value.date())
            elif isinstance(value, date):
                values.add(value)
        return values

    def _get_soll_duration(self, workbook, weekday: int) -> timedelta:
        if weekday >= 5:
            return timedelta(0)
        allgemein = workbook["Allgemein"]
        value = allgemein[f"C{26 + weekday}"].value
        normalized = normalize_excel_time(value)
        if normalized is None:
            raise ValueError(
                f"Sollarbeitszeit fuer Wochentag {weekday} konnte nicht gelesen werden."
            )
        return normalized
